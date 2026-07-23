import datetime
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Strip characters illegal in Excel worksheets (control chars except \t \n \r)
ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")  

def safe_str(val):
    if val is None:
        return ""
    s = str(val)
    return ILLEGAL_CHARS_RE.sub("", s)


COLORS = {
    "indigo": "4F46E5",
    "navy": "312E81",
    "green_fill": "DCFCE7",
    "green_font": "15803D",
    "slate_light": "F8FAFC",
    "gray_kpi": "EEF2F6",
    "border": "CBD5E1",
    "text_dark": "1E293B",
    "text_body": "334155",
    "white": "FFFFFF",
}


def make_styles():
    thin = Side(style="thin", color=COLORS["border"])
    return {
        "title": Font(name="Segoe UI", size=15, bold=True, color=COLORS["white"]),
        "section": Font(name="Segoe UI", size=11, bold=True, color=COLORS["text_dark"]),
        "header": Font(name="Segoe UI", size=10, bold=True, color=COLORS["white"]),
        "bold": Font(name="Segoe UI", size=10, bold=True, color=COLORS["text_dark"]),
        "regular": Font(name="Segoe UI", size=10, color=COLORS["text_body"]),
        "kpi": Font(name="Segoe UI", size=18, bold=True, color=COLORS["indigo"]),
        "pass_font": Font(name="Segoe UI", size=10, bold=True, color=COLORS["green_font"]),
        "fail_font": Font(name="Segoe UI", size=10, bold=True, color="DC2626"),
        "fill_title": PatternFill(start_color=COLORS["indigo"], end_color=COLORS["indigo"], fill_type="solid"),
        "fill_header": PatternFill(start_color=COLORS["navy"], end_color=COLORS["navy"], fill_type="solid"),
        "fill_pass": PatternFill(start_color=COLORS["green_fill"], end_color=COLORS["green_fill"], fill_type="solid"),
        "fill_kpi": PatternFill(start_color=COLORS["gray_kpi"], end_color=COLORS["gray_kpi"], fill_type="solid"),
        "fill_zebra": PatternFill(start_color=COLORS["slate_light"], end_color=COLORS["slate_light"], fill_type="solid"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


def generate_excel_report(
    test_results,
    suite_name,
    suite_description,
    output_path,
):
    """
    Generic Excel report generator.
    test_results: list of dicts with keys:
      id, module, name, preconditions, steps (list), expected, actual, status, execution_time, timestamp
    """
    print(f"[Reporter] Generating '{suite_name}' report → {len(test_results)} test cases …")
    st = make_styles()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary Dashboard ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary Dashboard"

    ws.merge_cells("A1:H2")
    ws["A1"].value = f"  Mind Mood AI ▸ {suite_name} — Automated Test Report"
    ws["A1"].font = st["title"]
    ws["A1"].fill = st["fill_title"]
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    meta = [
        ("A4", "Suite:", "B4", suite_description),
        ("A5", "Execution Date:", "B5", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("D4", "Environment:", "E4", "GitHub Actions CI / Vercel Production"),
        ("D5", "OS:", "E5", "Ubuntu-latest (Linux x64)"),
    ]
    for label_cell, label, val_cell, val in meta:
        ws[label_cell].value = label
        ws[label_cell].font = st["bold"]
        ws[val_cell].value = val
        ws[val_cell].font = st["regular"]

    total = len(test_results)
    passed = sum(1 for r in test_results if r.get("status", "Pass") == "Pass")
    failed = total - passed
    rate = f"{(passed / total * 100):.1f}%" if total else "0%"

    # KPI boxes
    def kpi_box(col_start, col_end, row_label, row_val, label, value, fill=None):
        ws.merge_cells(f"{col_start}{row_label}:{col_end}{row_label}")
        ws[f"{col_start}{row_label}"].value = label
        ws[f"{col_start}{row_label}"].font = st["bold"]
        ws[f"{col_start}{row_label}"].fill = fill or st["fill_kpi"]
        ws[f"{col_start}{row_label}"].alignment = st["center"]
        ws.merge_cells(f"{col_start}{row_val}:{col_end}{int(row_val)+1}")
        ws[f"{col_start}{row_val}"].value = value
        ws[f"{col_start}{row_val}"].font = st["kpi"]
        ws[f"{col_start}{row_val}"].fill = fill or st["fill_kpi"]
        ws[f"{col_start}{row_val}"].alignment = st["center"]

    kpi_box("A","B", 8, 9, "TOTAL TESTS", str(total))
    kpi_box("D","E", 8, 9, "PASSED", str(passed))
    kpi_box("G","H", 8, 9, "PASS RATE", rate, st["fill_pass"])

    # Module breakdown
    ws["A13"].value = "Module Breakdown"
    ws["A13"].font = st["section"]

    hdrs = ["Module / Category", "Test Cases", "Passed", "Failed", "Success Rate"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=14, column=ci)
        c.value = h; c.font = st["header"]; c.fill = st["fill_header"]
        c.alignment = st["center"]; c.border = st["border"]

    modules = {}
    for r in test_results:
        m = r.get("module", "General")
        modules.setdefault(m, []).append(r)

    row = 15
    for mod, cases in modules.items():
        p = sum(1 for r in cases if r.get("status") == "Pass")
        f = len(cases) - p
        cells_data = [mod, len(cases), p, f, f"{(p/len(cases)*100):.0f}%"]
        for ci, val in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=ci)
            cell.value = val
            cell.border = st["border"]
            cell.alignment = st["center"] if ci > 1 else st["left"]
            if row % 2 == 0:
                cell.fill = st["fill_zebra"]
            if ci == 5:
                cell.font = st["pass_font"]
                cell.fill = st["fill_pass"]
        row += 1

    # Total row
    for ci, val in enumerate(["TOTAL", total, passed, failed, rate], 1):
        cell = ws.cell(row=row, column=ci)
        cell.value = val; cell.font = st["bold"]; cell.border = st["border"]
        cell.alignment = st["center"] if ci > 1 else st["left"]
    ws.cell(row=row, column=5).fill = st["fill_pass"]
    ws.cell(row=row, column=5).font = Font(name="Segoe UI", size=10, bold=True, color=COLORS["white"])
    ws.cell(row=row, column=5).fill = PatternFill(start_color=COLORS["green_font"], end_color=COLORS["green_font"], fill_type="solid")

    # Resize columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value and c.row > 2), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 2: Full Test Details ───────────────────────────────────────────
    wd = wb.create_sheet(title="Test Execution Details")
    col_headers = [
        "Test ID", "Module", "Test Case Name", "Preconditions",
        "Test Steps", "Expected Result", "Actual Result",
        "Status", "Exec Time", "Timestamp"
    ]
    col_widths = [14, 22, 38, 36, 52, 40, 40, 10, 12, 20]

    for ci, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        c = wd.cell(row=1, column=ci)
        c.value = h; c.font = st["header"]; c.fill = st["fill_header"]
        c.alignment = st["center"]; c.border = st["border"]
        wd.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(test_results, 2):
        steps_str = "\n".join(r.get("steps", []))
        row_vals = [
            r.get("id", ""),
            r.get("module", ""),
            r.get("name", ""),
            r.get("preconditions", ""),
            steps_str,
            r.get("expected", ""),
            r.get("actual", ""),
            r.get("status", "Pass"),
            r.get("execution_time", "—"),
            r.get("timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = wd.cell(row=ri, column=ci)
            cell.value = safe_str(val)
            cell.font = st["regular"]
            cell.border = st["border"]
            cell.alignment = st["center"] if ci in (1, 8, 9, 10) else st["left"]
            if ri % 2 == 0 and ci != 8:
                cell.fill = st["fill_zebra"]
        # colour status column
        sc = wd.cell(row=ri, column=8)
        if sc.value == "Pass":
            sc.font = st["pass_font"]; sc.fill = st["fill_pass"]
        else:
            sc.font = st["fail_font"]
            sc.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    try:
        wb.save(output_path)
        print(f"[Reporter] ✓ Saved → {os.path.abspath(output_path)}")
    except PermissionError:
        base, ext = os.path.splitext(output_path)
        alt = f"{base}_alt{ext}"
        wb.save(alt)
        print(f"[Reporter] ✓ Saved (alt) → {os.path.abspath(alt)}")
